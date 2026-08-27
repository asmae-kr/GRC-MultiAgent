"""
Génère le rapport Excel final à partir de VOTRE template existant.
Le template de départ est VIDE (pas de données) — ce script réécrit donc
CHAQUE colonne (A à J) pour chaque ligne, en combinant :
  - les données d'origine du questionnaire (colonnes A à G, venant de Power Automate)
  - les résultats de l'analyse IA (colonnes H à J, venant des agents)
Ligne 1 = titres (déjà dans le template, jamais touchée).
Ligne 2 = question n°1, ligne 3 = question n°2, etc.
"""
import shutil
import openpyxl
 
 
def _find_row_by_label(ws, label: str) -> int:
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip() == label:
            return cell.row
    return None
 
 
def generate_excel_from_template(template_path: str, inner: dict, result: dict, output_path: str):
    shutil.copyfile(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
 
    report = result["report"]
    cover = {c["Champ"]: c.get("Valeur", "") for c in inner.get("cover", [])}
    questionnaire = inner.get("questionnaire", [])
    compliance_results = report["compliance_results"]
 
    # ── Onglet Cover ─────────────────────────────────────────────────
    ws_cover = wb["Cover"]
    for label, valeur in cover.items():
        row = _find_row_by_label(ws_cover, label)
        if row:
            ws_cover.cell(row=row, column=2, value=valeur)
 
    # ── Onglet Questionnaire : TOUTES les colonnes, A à J ──────────────
    ws_q = wb["Questionnaire"]
    # On avance en même temps dans les deux listes : question d'origine (A-G)
    # et son résultat d'analyse (H-J), grâce à zip() qui les associe une par une.
    for i, (q_row, comp) in enumerate(zip(questionnaire, compliance_results), start=2):
        ws_q.cell(row=i, column=1, value=q_row.get("No_x002e_", ""))                          # A - No.
        ws_q.cell(row=i, column=2, value=q_row.get("Domain", ""))                             # B - Domain
        ws_q.cell(row=i, column=3, value=q_row.get("Requirement", ""))                        # C - Requirement
        ws_q.cell(row=i, column=4, value=q_row.get("Question / Description", ""))             # D - Question
        ws_q.cell(row=i, column=5, value=q_row.get("Criticality", ""))                        # E - Criticality
        ws_q.cell(row=i, column=6, value=q_row.get("Response (Y/N/P/NA)", ""))                # F - Response
        ws_q.cell(row=i, column=7, value=q_row.get("Comments / Evidence reference", ""))      # G - Comments
        ws_q.cell(row=i, column=8, value=comp["score"])                                       # H - Score
        ws_q.cell(row=i, column=9, value=comp["justification"])                               # I - Justification
        follow_ups = comp.get("follow_up_questions", [])
        ws_q.cell(row=i, column=10, value="; ".join(follow_ups))                              # J - Questions
 
    wb.save(output_path)